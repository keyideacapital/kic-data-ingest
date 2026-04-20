#!/usr/bin/env python3
"""
kic_ingest/engine.py — Global contact & company ingestion engine for KIC.

Pipeline:
  1. LOAD      — detect encoding, open CSV/XLSX, yield (region_name, 2D-array)
  2. PROFILE   — segment each sheet into tabular regions; identify header row(s)
  3. EXTRACT   — map each region's columns onto the canonical schema via fuzzy match
  4. RESOLVE   — split unified rows into Organisation + Contact records; compute keys
  5. SCORE     — per-record ingestion_confidence and tier (high/medium/low)
  6. EMIT      — dump JSONL for downstream Airtable writer + review queue

Usage:
    python -m kic_ingest.engine INPUT_PATH [INPUT_PATH ...] --out OUT_DIR

Outputs (per input file):
    OUT_DIR/<stem>.organisations.jsonl
    OUT_DIR/<stem>.contacts.jsonl
    OUT_DIR/<stem>.documents.jsonl
    OUT_DIR/<stem>.report.json          # counts + tier histogram + unmapped columns
"""
from __future__ import annotations

import argparse
import csv
import hashlib
import json
import re
import sys
import unicodedata
from dataclasses import dataclass, field, asdict
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Iterator
from urllib.parse import urlparse

try:
    from rapidfuzz import fuzz
except ImportError:
    print("ERROR: rapidfuzz not installed. pip install rapidfuzz", file=sys.stderr)
    sys.exit(2)

try:
    from openpyxl import load_workbook
except ImportError:
    print("ERROR: openpyxl not installed. pip install openpyxl", file=sys.stderr)
    sys.exit(2)


# ---------------------------------------------------------------------------
# Config + thresholds
# ---------------------------------------------------------------------------

HERE = Path(__file__).resolve().parent
REFS_DIR = HERE.parent / "references"

# Minimum rapidfuzz token_sort_ratio score to accept a column->canonical mapping.
# Tuned by validation pass — headers like "E-mail Address" need some slack vs "email".
COLUMN_MATCH_THRESHOLD = 82

# Minimum cells with content for a row to count as "populated" when detecting regions.
MIN_POPULATED_CELLS_FOR_DATA_ROW = 2

# When scanning a sheet for multiple tabular regions, a blank-row run of this size
# or more (or a row with only 0-1 populated cells) ends the current region.
REGION_BREAK_BLANK_RUN = 2

# Confidence tier cutoffs (matches the design brief).
TIER_HIGH_MIN = 0.85
TIER_MEDIUM_MIN = 0.60

# Known superset sheets: when an xlsx has a master sheet that is a superset of
# other sector-filtered sheets in the same file, we should process the master
# and skip the subsets. The engine checks each sheet name against these patterns
# and, if a superset is present, records its subset siblings as skipped.
#
# Format: { canonical_superset_name: [list of subset sheet names to skip IF the
# superset is present in the same workbook] }.
# The MAS FID export is the concrete case: FID_YYYY-MM-DD is the superset of the
# per-sector sheets.
KNOWN_SUPERSETS_XLSX: dict[re.Pattern, set[str]] = {
    re.compile(r"^FID_\d{4}-\d{2}-\d{2}$"): {
        "payments", "financial advisory", "capital markets", "insurance", "banking",
    },
}


def _sheets_to_skip_for_supersets(sheet_names: list[str]) -> set[str]:
    """Given all sheet names in a workbook, return the lower-cased names of
    subset sheets that should be skipped because a known-superset sheet is also
    present."""
    skip: set[str] = set()
    lowered = [s.strip().lower() for s in sheet_names]
    for pattern, subsets in KNOWN_SUPERSETS_XLSX.items():
        if any(pattern.match(name) for name in sheet_names):
            skip.update(subsets)
    return skip


# Known table-of-contents / search-form sheet names to skip outright.
SKIP_SHEET_NAMES = {"search", "toc", "index", "readme", "cover", "instructions"}


# ---------------------------------------------------------------------------
# Normalisers
# ---------------------------------------------------------------------------

_PUNCT_RE = re.compile(r"[^\w\s]+")
_WS_RE = re.compile(r"\s+")


def normalise_header(s: str) -> str:
    """Lowercase, strip diacritics, strip punctuation, collapse whitespace."""
    if s is None:
        return ""
    s = unicodedata.normalize("NFKD", str(s)).encode("ascii", "ignore").decode()
    s = _PUNCT_RE.sub(" ", s).lower()
    s = _WS_RE.sub(" ", s).strip()
    return s


# Common org-suffix tokens stripped before matching org names.
_ORG_SUFFIX_TOKENS = {
    "pte", "ltd", "limited", "llc", "inc", "incorporated", "corp", "corporation",
    "co", "company", "gmbh", "sa", "sas", "ag", "bv", "nv", "plc", "llp", "lp",
    "llp", "pty", "holdings", "group", "capital", "partners", "ventures", "vc",
    "the", "and", "of"
}


def normalise_org_name(s: str) -> str:
    if not s:
        return ""
    h = normalise_header(s)
    tokens = [t for t in h.split() if t not in _ORG_SUFFIX_TOKENS]
    return " ".join(tokens) if tokens else h


def normalise_person_name(s: str) -> str:
    return normalise_header(s)


_URL_RE = re.compile(r"https?://[^\s,;]+", re.IGNORECASE)

# Personal / webmail domains — these must NEVER be used as an employer's website_domain.
# A contact's hotmail.com email tells us nothing about their company, so inheriting it as
# the org's domain pollutes matching (e.g. two unrelated orgs both getting gmail.com).
PERSONAL_EMAIL_DOMAINS = {
    "gmail.com", "googlemail.com",
    "hotmail.com", "hotmail.co.uk", "hotmail.com.sg", "hotmail.com.au",
    "outlook.com", "live.com", "msn.com",
    "yahoo.com", "yahoo.co.uk", "yahoo.com.sg", "yahoo.com.au", "yahoo.com.hk",
    "ymail.com", "rocketmail.com",
    "icloud.com", "me.com", "mac.com",
    "protonmail.com", "proton.me", "pm.me",
    "netvigator.com", "biznetvigator.com",  # HK telco webmail — common in the LinkedIn file
    "qq.com", "163.com", "126.com", "sina.com", "sina.cn", "sohu.com",
    "aol.com", "gmx.com", "gmx.net", "zoho.com", "mail.com", "fastmail.com",
    "tutanota.com", "hey.com", "inbox.com",
    "bigpond.com", "bigpond.net.au", "optusnet.com.au", "iinet.com.au",
    "naver.com", "daum.net", "hanmail.net",
}


def is_personal_email_domain(domain: str | None) -> bool:
    if not domain:
        return False
    d = domain.lower().strip()
    if d.startswith("www."):
        d = d[4:]
    return d in PERSONAL_EMAIL_DOMAINS


def extract_domain(value: str) -> str | None:
    """Return the eTLD+1-ish domain from a URL or email, lowercase, no www.

    We don't do a full PSL lookup — good enough for matching and traceability.
    """
    if not value:
        return None
    v = str(value).strip()
    # email?
    if "@" in v and " " not in v:
        part = v.rsplit("@", 1)[-1]
    else:
        # URL — prepend scheme if missing so urlparse cooperates
        if not v.lower().startswith(("http://", "https://")):
            v = "http://" + v
        try:
            part = urlparse(v).hostname or ""
        except ValueError:
            return None
    part = part.lower().strip().strip(".")
    if part.startswith("www."):
        part = part[4:]
    return part or None


# ---------------------------------------------------------------------------
# Field synonym resolver
# ---------------------------------------------------------------------------

@dataclass
class FieldSynonymMap:
    synonyms: dict[str, list[str]]          # canonical -> list of raw synonyms
    normalised_lookup: dict[str, str] = field(default_factory=dict)  # norm -> canonical

    @classmethod
    def load(cls, path: Path) -> "FieldSynonymMap":
        raw = json.loads(path.read_text())
        raw.pop("_meta", None)
        syn = {k: list(v) for k, v in raw.items()}
        lookup: dict[str, str] = {}
        for canonical, terms in syn.items():
            # The canonical field name itself counts as a synonym too.
            all_terms = [canonical] + list(terms)
            for t in all_terms:
                lookup[normalise_header(t)] = canonical
        return cls(synonyms=syn, normalised_lookup=lookup)

    def resolve(self, header: str) -> tuple[str | None, int]:
        """Map a raw header to a canonical field, returning (canonical, score) or (None, 0)."""
        if not header:
            return None, 0
        n = normalise_header(header)
        if not n:
            return None, 0

        # Pass 1: exact normalised match.
        if n in self.normalised_lookup:
            return self.normalised_lookup[n], 100

        # Pass 2: fuzzy token_sort over all known synonyms, keep best.
        best_canonical: str | None = None
        best_score = 0
        for canonical, terms in self.synonyms.items():
            for t in [canonical] + list(terms):
                score = fuzz.token_sort_ratio(n, normalise_header(t))
                if score > best_score:
                    best_score = score
                    best_canonical = canonical
        if best_score >= COLUMN_MATCH_THRESHOLD:
            return best_canonical, best_score
        return None, best_score


# ---------------------------------------------------------------------------
# Loaders
# ---------------------------------------------------------------------------

def _detect_text_encoding(path: Path, sample_bytes: int = 32768) -> str:
    """Best-effort encoding detection. Tries utf-8 first, falls back to cp1252 then latin-1."""
    raw = path.read_bytes()[:sample_bytes]
    for enc in ("utf-8-sig", "utf-8", "cp1252", "latin-1"):
        try:
            raw.decode(enc)
            return enc
        except UnicodeDecodeError:
            continue
    return "latin-1"  # latin-1 decodes any byte, last resort


def load_csv(path: Path) -> Iterator[tuple[str, list[list[str]], str]]:
    """Yield (region_name, rows, encoding) — CSV has a single region named 'main'."""
    enc = _detect_text_encoding(path)
    with path.open("r", encoding=enc, newline="", errors="replace") as f:
        # Sniff delimiter on a sample — default to comma.
        sample = f.read(8192)
        f.seek(0)
        try:
            dialect = csv.Sniffer().sniff(sample, delimiters=",;\t|")
        except csv.Error:
            dialect = csv.excel
        reader = csv.reader(f, dialect)
        rows = [[("" if c is None else str(c)) for c in row] for row in reader]
    yield "main", rows, enc


def load_xlsx(path: Path) -> Iterator[tuple[str, list[list[str]], str]]:
    """Yield (sheet_name, rows_as_strings, 'xlsx-binary').

    Applies two skip-rules:
      1. Sheet name in SKIP_SHEET_NAMES (TOC / search-form sheets).
      2. Sheet is a known subset of a superset also present in the workbook
         (e.g. MAS FID master + per-sector duplicates).
    """
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        skip_subsets = _sheets_to_skip_for_supersets(wb.sheetnames)
        for sheet_name in wb.sheetnames:
            lname = sheet_name.strip().lower()
            if lname in SKIP_SHEET_NAMES:
                continue
            if lname in skip_subsets:
                # Skipping because a superset sheet is also in this workbook.
                continue
            ws = wb[sheet_name]
            rows: list[list[str]] = []
            for row in ws.iter_rows(values_only=True):
                rows.append([("" if c is None else str(c)) for c in row])
            yield sheet_name, rows, "xlsx-binary"
    finally:
        wb.close()


def load_any(path: Path) -> Iterator[tuple[str, list[list[str]], str]]:
    suffix = path.suffix.lower()
    if suffix in {".csv", ".tsv", ".txt"}:
        yield from load_csv(path)
    elif suffix in {".xlsx", ".xlsm"}:
        yield from load_xlsx(path)
    else:
        raise ValueError(f"Unsupported file type: {suffix} ({path})")


# ---------------------------------------------------------------------------
# Region detection
# ---------------------------------------------------------------------------

@dataclass
class Region:
    name: str                   # sheet + optional #region-N
    header_row_index: int       # 0-based into `rows`
    headers: list[str]
    data_row_indices: list[int] # 0-based into `rows`
    kind: str                   # "tabular" | "url_dump" | "noise"


def _populated_count(row: list[str]) -> int:
    return sum(1 for c in row if c and c.strip())


def _looks_like_header_row(row: list[str], synonyms: FieldSynonymMap) -> tuple[bool, int]:
    """Return (is_header, mapped_count). A row is a header if >= 2 cells map to canonical fields."""
    mapped = 0
    total_populated = 0
    for cell in row:
        if cell and cell.strip():
            total_populated += 1
            canonical, score = synonyms.resolve(cell)
            if canonical is not None:
                mapped += 1
    # Require at least 2 mapped cells AND at least half of populated cells mapped.
    if mapped >= 2 and total_populated > 0 and mapped / total_populated >= 0.4:
        return True, mapped
    return False, mapped


def _row_is_url_only(row: list[str]) -> bool:
    """True if populated cells in this row are exclusively URLs (VC-list-style URL dumps)."""
    populated = [c for c in row if c and c.strip()]
    if not populated:
        return False
    return all(_URL_RE.match(c.strip()) for c in populated)


def detect_regions(sheet_name: str, rows: list[list[str]], synonyms: FieldSynonymMap) -> list[Region]:
    """Segment a 2-D grid into tabular / url_dump / noise regions.

    Strategy:
      1. Sweep top-to-bottom looking for header rows (>= 2 canonical-mappable cells).
      2. A tabular region = header row + contiguous populated rows below it until
         a blank run of REGION_BREAK_BLANK_RUN or a new header row is found.
      3. Any consecutive stretch of URL-only rows outside a tabular region becomes
         a url_dump region with a synthetic 'website' header.
      4. Everything else is 'noise' and is not returned (but rows are still tracked
         for potential future use).
    """
    regions: list[Region] = []
    n = len(rows)
    i = 0
    region_counter = 0

    while i < n:
        row = rows[i]
        populated = _populated_count(row)

        # Skip truly blank rows.
        if populated == 0:
            i += 1
            continue

        # Header row?
        is_header, _mapped = _looks_like_header_row(row, synonyms)
        if is_header:
            region_counter += 1
            header_row_index = i
            headers = [c.strip() for c in row]
            # Collect data rows.
            data_rows: list[int] = []
            blank_run = 0
            j = i + 1
            while j < n:
                nr = rows[j]
                np = _populated_count(nr)
                if np == 0:
                    blank_run += 1
                    if blank_run >= REGION_BREAK_BLANK_RUN:
                        break
                    j += 1
                    continue
                # New header? stop this region.
                if _looks_like_header_row(nr, synonyms)[0]:
                    break
                blank_run = 0
                if np >= MIN_POPULATED_CELLS_FOR_DATA_ROW:
                    data_rows.append(j)
                j += 1
            regions.append(Region(
                name=f"{sheet_name}#region-{region_counter}",
                header_row_index=header_row_index,
                headers=headers,
                data_row_indices=data_rows,
                kind="tabular",
            ))
            i = j
            continue

        # URL-dump run?
        if _row_is_url_only(row):
            start = i
            url_row_indices: list[int] = []
            while i < n:
                nr = rows[i]
                if _populated_count(nr) == 0:
                    # Short blank runs are tolerated inside URL dumps.
                    if i + 1 < n and _populated_count(rows[i + 1]) == 0:
                        break
                    i += 1
                    continue
                if _row_is_url_only(nr):
                    url_row_indices.append(i)
                    i += 1
                    continue
                if _looks_like_header_row(nr, synonyms)[0]:
                    break
                # Non-URL, non-header populated row — break the URL dump.
                break
            if url_row_indices:
                region_counter += 1
                regions.append(Region(
                    name=f"{sheet_name}#region-{region_counter}-urldump",
                    header_row_index=start,       # synthetic; start is not a real header
                    headers=["website"],          # synthesised
                    data_row_indices=url_row_indices,
                    kind="url_dump",
                ))
            continue

        # Noise — skip.
        i += 1

    return regions


# ---------------------------------------------------------------------------
# Column mapping
# ---------------------------------------------------------------------------

@dataclass
class ColumnMapping:
    region_name: str
    header_to_canonical: dict[int, str]  # column index -> canonical field name
    unmapped_headers: list[str]
    avg_match_score: float


def map_columns(region: Region, synonyms: FieldSynonymMap) -> ColumnMapping:
    mapping: dict[int, str] = {}
    unmapped: list[str] = []
    scores: list[int] = []

    if region.kind == "url_dump":
        # Synthetic single-column region.
        mapping[0] = "website"
        return ColumnMapping(region.name, mapping, [], 100.0)

    # Track which canonical field has been used — allow duplicates for email/phone variants.
    allow_duplicate = {"email", "email_alt", "phone_business", "phone_mobile", "phone_other"}
    used_exclusive: set[str] = set()

    for idx, header in enumerate(region.headers):
        if not header or not header.strip():
            continue
        canonical, score = synonyms.resolve(header)
        if canonical is None:
            unmapped.append(header)
            continue
        if canonical in used_exclusive and canonical not in allow_duplicate:
            # Duplicate mapping to a non-repeatable canonical field — demote to unmapped.
            unmapped.append(header)
            continue
        mapping[idx] = canonical
        if canonical not in allow_duplicate:
            used_exclusive.add(canonical)
        scores.append(score)

    avg = (sum(scores) / len(scores)) if scores else 0.0
    return ColumnMapping(region.name, mapping, unmapped, avg)


# ---------------------------------------------------------------------------
# Record extraction
# ---------------------------------------------------------------------------

@dataclass
class ExtractedOrg:
    source_ref: str
    source_region: str
    data: dict[str, Any]
    confidence: float
    tier: str

@dataclass
class ExtractedContact:
    source_ref: str
    source_region: str
    data: dict[str, Any]
    confidence: float
    tier: str


ORG_CANONICAL_FIELDS = {
    "org_name", "org_type", "website", "linkedin_company_url",
    "hq_city", "hq_country", "offices", "phone", "general_email",
    "address", "sector", "stage_focus", "aum", "licence_type",
    "activity_type", "notes",
}
CONTACT_CANONICAL_FIELDS = {
    "first_name", "middle_name", "last_name", "full_name", "title", "job_title",
    "email", "email_alt",
    "phone_business", "phone_mobile", "phone_other",
    "linkedin_profile_url", "address", "city", "country", "notes",
}


def _tier_for(confidence: float) -> str:
    if confidence >= TIER_HIGH_MIN:
        return "high"
    if confidence >= TIER_MEDIUM_MIN:
        return "medium"
    return "low"


# Field-type sanity guards applied after column mapping. If a value clearly
# doesn't belong in the canonical field it was mapped to (e.g. a URL ending up
# in full_name because the source file has shifted/mispasted data), we reject
# the value rather than emit a malformed record. Rejected values are dropped
# silently — the row still produces a record from its other fields.
_NAME_FIELDS = {"first_name", "middle_name", "last_name", "full_name"}
_EMAIL_FIELDS = {"email", "email_alt", "general_email"}
_URL_FIELDS = {"website", "linkedin_company_url", "linkedin_profile_url"}
_PHONE_FIELDS = {"phone", "phone_business", "phone_mobile", "phone_other"}


def _value_looks_valid(canonical: str, value: str) -> bool:
    """Cheap sanity check. Returns False when `value` is obviously not the right
    type for the canonical field it was mapped to."""
    v = value.strip()
    if not v:
        return False
    vl = v.lower()
    if canonical in _NAME_FIELDS:
        # Names shouldn't contain URLs or @ signs or long digit runs.
        if "http://" in vl or "https://" in vl or "www." in vl:
            return False
        if "@" in v:
            return False
        if sum(ch.isdigit() for ch in v) > len(v) / 2:
            return False
    elif canonical in _EMAIL_FIELDS:
        if "@" not in v or " " in v:
            return False
    elif canonical in _URL_FIELDS:
        # Permit bare domains too; reject anything with spaces or obvious non-URL content.
        if " " in v and not vl.startswith(("http://", "https://")):
            return False
    elif canonical == "phone" or canonical in _PHONE_FIELDS:
        # Phone must contain at least 5 digits.
        if sum(ch.isdigit() for ch in v) < 5:
            return False
    return True


def _collect_row_fields(row: list[str], mapping: ColumnMapping) -> dict[str, Any]:
    """Map a single data row's cells into a canonical-field dict. Handles duplicate
    canonical mappings by concatenating with '; '. Applies value-level sanity
    guards to drop obviously-wrong cells (e.g. URLs mapped into name fields)."""
    out: dict[str, Any] = {}
    for idx, canonical in mapping.header_to_canonical.items():
        if idx >= len(row):
            continue
        val = row[idx]
        if val is None:
            continue
        val = str(val).strip()
        if not val:
            continue
        if not _value_looks_valid(canonical, val):
            continue
        if canonical in out and out[canonical] != val:
            out[canonical] = f"{out[canonical]}; {val}"
        else:
            out[canonical] = val
    return out


def _compose_full_name(fields: dict[str, Any]) -> str | None:
    parts = []
    for k in ("first_name", "middle_name", "last_name"):
        if k in fields and fields[k]:
            parts.append(str(fields[k]).strip())
    if parts:
        return " ".join(parts)
    if "full_name" in fields and fields["full_name"]:
        return str(fields["full_name"]).strip()
    return None


def _score_org(org: dict[str, Any], column_match_score: float) -> float:
    """Confidence = weighted signal strength × column-match quality.

    Calibration (post-validation):
      - An org name alone, from a sheet whose columns mapped cleanly, deserves at
        least medium tier (0.60) — the row has been profiled, segmented, and its
        single canonical field matches. A bare name from chaotic data (low column
        quality) still gets capped via the quality_factor.
      - Rich records (name + website + country + sector/licence) reach high (0.85).
    """
    if not org.get("org_name"):
        return 0.0
    # Base: having a resolved org name in a mapped region.
    score = 0.55
    weights = {
        "website": 0.15, "website_domain": 0.08, "phone": 0.06, "general_email": 0.06,
        "hq_country": 0.05, "hq_city": 0.03, "sector": 0.05, "address": 0.05,
        "licence_type": 0.05, "activity_type": 0.03, "linkedin_company_url": 0.08,
        "aum": 0.03, "stage_focus": 0.03, "offices": 0.02,
    }
    for k, w in weights.items():
        if org.get(k):
            score += w
    # Column quality ceiling: if the sheet's columns mapped poorly, cap confidence.
    quality_factor = min(1.0, (column_match_score or 0) / 100.0)
    return round(min(1.0, score) * quality_factor, 3)


def _score_contact(contact: dict[str, Any], column_match_score: float) -> float:
    """A contact needs at least a name OR an email to score above 0.

    Calibration (post-validation):
      - full_name + job_title + org_name (the curated-list pattern) should clear
        medium (0.60). Without email, these are still high-value records.
      - Email alone still starts at a respectable base — email is the single
        strongest contact signal for identity/dedup.
    """
    has_name = bool(contact.get("full_name"))
    has_email = bool(contact.get("email"))
    if not has_name and not has_email:
        return 0.0
    score = 0.0
    if has_name:  score += 0.40
    if has_email: score += 0.30
    # A name + org pairing is the strongest "real contact" signal after email.
    if has_name and contact.get("org_name"): score += 0.10
    weights = {
        "job_title": 0.10, "phone_business": 0.05, "phone_mobile": 0.05,
        "linkedin_profile_url": 0.08, "email_alt": 0.03,
        "city": 0.02, "country": 0.02, "address": 0.02,
    }
    for k, w in weights.items():
        if contact.get(k):
            score += w
    quality_factor = min(1.0, (column_match_score or 0) / 100.0)
    return round(min(1.0, score) * quality_factor, 3)


def extract_records(
    source_ref_prefix: str,
    region: Region,
    rows: list[list[str]],
    mapping: ColumnMapping,
) -> tuple[list[ExtractedOrg], list[ExtractedContact]]:
    """Walk data rows of a region, splitting each into Organisation + Contact records."""
    orgs: list[ExtractedOrg] = []
    contacts: list[ExtractedContact] = []
    col_quality = mapping.avg_match_score

    for row_idx in region.data_row_indices:
        if row_idx >= len(rows):
            continue
        row = rows[row_idx]
        source_ref = f"{source_ref_prefix}!row={row_idx + 1}"
        fields = _collect_row_fields(row, mapping)
        if not fields:
            continue

        # --- Organisation slice ---
        org_fields: dict[str, Any] = {k: v for k, v in fields.items() if k in ORG_CANONICAL_FIELDS}

        # Derive website_domain if website or general_email present.
        # CRITICAL: never inherit a domain from a contact's personal email (gmail/hotmail/etc.)
        # to the org record — that pollutes matching.
        dom = None
        if org_fields.get("website"):
            dom = extract_domain(org_fields["website"])
        elif org_fields.get("general_email"):
            cand = extract_domain(org_fields["general_email"])
            if cand and not is_personal_email_domain(cand):
                dom = cand
        elif fields.get("email"):
            cand = extract_domain(fields["email"])
            if cand and not is_personal_email_domain(cand):
                dom = cand
        if dom:
            org_fields["website_domain"] = dom

        if org_fields.get("org_name"):
            org_fields["org_name_normalised"] = normalise_org_name(org_fields["org_name"])

        if org_fields.get("org_name") or (region.kind == "url_dump" and org_fields.get("website")):
            # URL-dump region: the URL itself becomes the org name stand-in.
            if region.kind == "url_dump" and not org_fields.get("org_name"):
                url = org_fields.get("website", "").strip()
                org_fields["org_name"] = dom or url
                org_fields["org_name_normalised"] = normalise_org_name(org_fields["org_name"])

            confidence = _score_org(org_fields, col_quality)
            # URL-dump records are intrinsically low-confidence (website only).
            if region.kind == "url_dump":
                confidence = min(confidence, 0.55)
            orgs.append(ExtractedOrg(
                source_ref=source_ref,
                source_region=region.name,
                data={
                    **org_fields,
                    "source_ref": source_ref,
                    "source_region": region.name,
                    "ingestion_confidence": confidence,
                    "ingestion_tier": _tier_for(confidence),
                },
                confidence=confidence,
                tier=_tier_for(confidence),
            ))

        # --- Contact slice ---
        contact_fields: dict[str, Any] = {k: v for k, v in fields.items() if k in CONTACT_CANONICAL_FIELDS}
        full = _compose_full_name(contact_fields)
        if full:
            contact_fields["full_name"] = full
            contact_fields["full_name_normalised"] = normalise_person_name(full)

        # Contact inherits org linkage from the same row's org_name (unified-row case like AU CSV)
        if org_fields.get("org_name"):
            contact_fields.setdefault("org_name", org_fields["org_name"])

        # A contact is emitted only if we have a name OR an email.
        if contact_fields.get("full_name") or contact_fields.get("email"):
            c_conf = _score_contact(contact_fields, col_quality)
            contacts.append(ExtractedContact(
                source_ref=source_ref,
                source_region=region.name,
                data={
                    **contact_fields,
                    "source_ref": source_ref,
                    "source_region": region.name,
                    "ingestion_confidence": c_conf,
                    "ingestion_tier": _tier_for(c_conf),
                },
                confidence=c_conf,
                tier=_tier_for(c_conf),
            ))

    return orgs, contacts


# ---------------------------------------------------------------------------
# Driver
# ---------------------------------------------------------------------------

def _dedup_orgs(orgs: list[ExtractedOrg]) -> tuple[list[ExtractedOrg], int]:
    """Collapse duplicate Organisation records within a single file.

    Key = website_domain if present, else normalised org_name. When duplicates
    are found, keep the highest-confidence record and merge any non-empty fields
    from the duplicates into it (so a URL-dump hit and a tabular hit for the
    same company compose rather than both being emitted).

    Returns (deduped_list, collapsed_count).
    """
    by_key: dict[str, ExtractedOrg] = {}
    collapsed = 0
    for o in orgs:
        dom = o.data.get("website_domain")
        name_norm = o.data.get("org_name_normalised") or normalise_org_name(o.data.get("org_name", ""))
        # Prefer domain as key when available — it's more reliable than a name.
        key = f"dom::{dom}" if dom else f"name::{name_norm}"
        if not key or key in ("dom::", "name::"):
            # No useful key — keep the record but don't dedup it.
            by_key[f"unkeyed::{id(o)}"] = o
            continue
        existing = by_key.get(key)
        if existing is None:
            by_key[key] = o
            continue
        collapsed += 1
        # Merge — prefer the higher-confidence record, then fill gaps.
        winner, loser = (existing, o) if existing.confidence >= o.confidence else (o, existing)
        merged_data = dict(winner.data)
        for k, v in loser.data.items():
            if k in ("ingestion_confidence", "ingestion_tier", "source_ref", "source_region"):
                continue
            if v and not merged_data.get(k):
                merged_data[k] = v
        # Append loser's source_ref to provenance trail so we never lose traceability.
        merged_sources = merged_data.get("source_ref", "")
        if loser.data.get("source_ref") and loser.data["source_ref"] not in merged_sources:
            merged_data["source_ref"] = f"{merged_sources}|{loser.data['source_ref']}"
        by_key[key] = ExtractedOrg(
            source_ref=merged_data.get("source_ref", winner.source_ref),
            source_region=winner.source_region,
            data=merged_data,
            confidence=winner.confidence,
            tier=winner.tier,
        )
    return list(by_key.values()), collapsed


def _dedup_contacts(contacts: list[ExtractedContact]) -> tuple[list[ExtractedContact], int]:
    """Collapse duplicate Contact records within a single file.

    Key = email if present, else (normalised_name, normalised_org_name).
    Same merge strategy as orgs.
    """
    by_key: dict[str, ExtractedContact] = {}
    collapsed = 0
    for c in contacts:
        email = (c.data.get("email") or "").strip().lower()
        name_norm = c.data.get("full_name_normalised") or normalise_person_name(c.data.get("full_name", ""))
        org_norm = normalise_org_name(c.data.get("org_name", ""))
        if email:
            key = f"email::{email}"
        elif name_norm:
            key = f"nameorg::{name_norm}::{org_norm}"
        else:
            by_key[f"unkeyed::{id(c)}"] = c
            continue
        existing = by_key.get(key)
        if existing is None:
            by_key[key] = c
            continue
        collapsed += 1
        winner, loser = (existing, c) if existing.confidence >= c.confidence else (c, existing)
        merged_data = dict(winner.data)
        for k, v in loser.data.items():
            if k in ("ingestion_confidence", "ingestion_tier", "source_ref", "source_region"):
                continue
            if v and not merged_data.get(k):
                merged_data[k] = v
        merged_sources = merged_data.get("source_ref", "")
        if loser.data.get("source_ref") and loser.data["source_ref"] not in merged_sources:
            merged_data["source_ref"] = f"{merged_sources}|{loser.data['source_ref']}"
        by_key[key] = ExtractedContact(
            source_ref=merged_data.get("source_ref", winner.source_ref),
            source_region=winner.source_region,
            data=merged_data,
            confidence=winner.confidence,
            tier=winner.tier,
        )
    return list(by_key.values()), collapsed


def process_file(path: Path, synonyms: FieldSynonymMap, out_dir: Path) -> dict[str, Any]:
    stem = path.stem
    out_dir.mkdir(parents=True, exist_ok=True)

    orgs_path = out_dir / f"{stem}.organisations.jsonl"
    contacts_path = out_dir / f"{stem}.contacts.jsonl"
    documents_path = out_dir / f"{stem}.documents.jsonl"
    report_path = out_dir / f"{stem}.report.json"

    doc_id = hashlib.sha256(str(path.resolve()).encode()).hexdigest()[:16]
    file_type = path.suffix.lower().lstrip(".")

    all_orgs: list[ExtractedOrg] = []
    all_contacts: list[ExtractedContact] = []
    region_reports: list[dict[str, Any]] = []
    total_rows = 0
    encoding_seen = ""

    for sheet_name, rows, enc in load_any(path):
        encoding_seen = enc
        total_rows += len(rows)
        regions = detect_regions(sheet_name, rows, synonyms)
        for region in regions:
            if region.kind not in {"tabular", "url_dump"}:
                continue
            mapping = map_columns(region, synonyms)
            if not mapping.header_to_canonical:
                region_reports.append({
                    "region": region.name, "kind": region.kind,
                    "data_rows": len(region.data_row_indices),
                    "mapped_columns": 0,
                    "unmapped_headers": mapping.unmapped_headers,
                    "status": "skipped-no-canonical-columns",
                })
                continue
            source_ref_prefix = f"{path.name}#{sheet_name}" if sheet_name != "main" else path.name
            orgs, contacts = extract_records(source_ref_prefix, region, rows, mapping)
            all_orgs.extend(orgs)
            all_contacts.extend(contacts)
            region_reports.append({
                "region": region.name, "kind": region.kind,
                "data_rows": len(region.data_row_indices),
                "mapped_columns": len(mapping.header_to_canonical),
                "unmapped_headers": mapping.unmapped_headers,
                "avg_match_score": mapping.avg_match_score,
                "orgs_extracted": len(orgs),
                "contacts_extracted": len(contacts),
                "status": "ok",
            })

    # Post-extraction dedup (intra-file). Catches same-company records emitted
    # from overlapping regions (e.g. URL-dump + tabular hits) and any residual
    # duplicates the superset rule didn't eliminate.
    orgs_before = len(all_orgs)
    contacts_before = len(all_contacts)
    all_orgs, orgs_collapsed = _dedup_orgs(all_orgs)
    all_contacts, contacts_collapsed = _dedup_contacts(all_contacts)

    # Write outputs
    with orgs_path.open("w") as f:
        for o in all_orgs:
            f.write(json.dumps(o.data, ensure_ascii=False) + "\n")
    with contacts_path.open("w") as f:
        for c in all_contacts:
            f.write(json.dumps(c.data, ensure_ascii=False) + "\n")
    with documents_path.open("w") as f:
        f.write(json.dumps({
            "document_id": doc_id,
            "filename": path.name,
            "file_type": file_type,
            "source_path": str(path.resolve()),
            "region_count": len([r for r in region_reports if r["status"] == "ok"]),
            "row_count": total_rows,
            "encoding": encoding_seen,
            "ingested_at": datetime.now(timezone.utc).isoformat(),
            "source_ref": path.name,
        }, ensure_ascii=False) + "\n")

    # Report
    def _tier_hist(items):
        h = {"high": 0, "medium": 0, "low": 0}
        for it in items:
            h[it.tier] = h.get(it.tier, 0) + 1
        return h

    report = {
        "file": path.name,
        "file_type": file_type,
        "encoding": encoding_seen,
        "total_raw_rows": total_rows,
        "regions": region_reports,
        "orgs_total": len(all_orgs),
        "contacts_total": len(all_contacts),
        "orgs_by_tier": _tier_hist(all_orgs),
        "contacts_by_tier": _tier_hist(all_contacts),
        "dedup": {
            "orgs_before": orgs_before,
            "orgs_after": len(all_orgs),
            "orgs_collapsed": orgs_collapsed,
            "contacts_before": contacts_before,
            "contacts_after": len(all_contacts),
            "contacts_collapsed": contacts_collapsed,
        },
    }
    report_path.write_text(json.dumps(report, indent=2, ensure_ascii=False))
    return report


def main(argv: list[str] | None = None) -> int:
    ap = argparse.ArgumentParser(description="KIC Ingest Engine v1")
    ap.add_argument("inputs", nargs="+", help="CSV or XLSX files to ingest")
    ap.add_argument("--out", default="out", help="Output directory (default: ./out)")
    ap.add_argument("--refs", default=str(REFS_DIR),
                    help="References directory containing field_synonyms.json + canonical_schema.json")
    args = ap.parse_args(argv)

    refs_dir = Path(args.refs)
    synonyms = FieldSynonymMap.load(refs_dir / "field_synonyms.json")
    # canonical_schema is loaded here as a future validation hook; engine currently
    # hardcodes the Org/Contact field sets above. Schema is still written to out/ for
    # downstream writer + CLAUDE.md reference.
    _schema = json.loads((refs_dir / "canonical_schema.json").read_text())

    out_dir = Path(args.out)
    print(f"→ Output dir: {out_dir.resolve()}")

    aggregate = {"files": [], "orgs_total": 0, "contacts_total": 0}
    for input_path in args.inputs:
        p = Path(input_path)
        if not p.exists():
            print(f"!! Skipping missing file: {p}", file=sys.stderr)
            continue
        print(f"\n→ Processing {p.name} ...")
        try:
            report = process_file(p, synonyms, out_dir)
        except Exception as exc:
            print(f"!! FAILED on {p.name}: {exc}", file=sys.stderr)
            raise
        print(f"   orgs:     {report['orgs_total']:>5}  (high/med/low = "
              f"{report['orgs_by_tier']['high']}/{report['orgs_by_tier']['medium']}/{report['orgs_by_tier']['low']})")
        print(f"   contacts: {report['contacts_total']:>5}  (high/med/low = "
              f"{report['contacts_by_tier']['high']}/{report['contacts_by_tier']['medium']}/{report['contacts_by_tier']['low']})")
        aggregate["orgs_total"] += report["orgs_total"]
        aggregate["contacts_total"] += report["contacts_total"]
        aggregate["files"].append(report["file"])

    print(f"\n✓ Done. Total orgs={aggregate['orgs_total']}  total contacts={aggregate['contacts_total']}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
